# Import necessary libraries
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import timedelta


def extract_sec(text):
    """
    Extracts the 'sec' part from a string.
    Assumes the 'sec' is the part before 'T'.
    """
    parts = text.split("T")
    return parts[0]


def extract_ter(text):
    """
    Extracts the 'ter' part from a string.
    Assumes the 'ter' is the part after 'T'.
    """
    parts = text.split("T")
    return "T" + parts[1]


def extract_date(date, hour):
    """
    Replaces the hour of a datetime object with a new hour.
    """
    return date.replace(hour=hour)


def calculate_hours_difference(start_date, end_date):
    """
    Calculates the difference in hours between two datetime objects.
    """
    time_difference = end_date - start_date
    difference_in_hours = time_difference.total_seconds() / 3600
    return difference_in_hours


def is_even(number):
    """
    Checks if a number is even.
    """
    if number % 2 == 0:
        return True
    else:
        return False


def sum_chunks(data_list, chunk_sizes):
    """
    Splits a list into chunks of specified sizes and calculates the average of each chunk.
    """
    result_list = []
    start_index = 0
    pattern_index = 0

    while start_index < len(data_list):
        chunk_size = chunk_sizes[pattern_index % len(chunk_sizes)]
        end_index = start_index + chunk_size

        # Ensures the last chunk doesn't go past the end of the list
        chunk = data_list[start_index:end_index]
        result_list.append(sum(chunk) / chunk_size)

        start_index = end_index
        pattern_index += 1

    return result_list


def xls_to_dict(original_path):
    """
    Reads an Excel file and converts its data into a structured dictionary.

    Args:
        original_path (str): The path to the input Excel file.

    Returns:
        tuple: A tuple containing the processed dictionary and the earliest start date.
    """
    # Load the entire workbook
    workbook = load_workbook(original_path)
    print("DEBUG: Workbook loaded successfully.")

    # Select the active sheet
    sheet = workbook.active

    # Initialize variables
    mv2 = {}
    TRD_start = None
    TRD_end = None

    # Iterate over rows to find the earliest TRD_start
    for row in sheet.iter_rows(
        min_row=2, max_row=10, max_col=11, min_col=10, values_only=True
    ):
        if row[0] is None:
            continue
        date = extract_date(row[0], row[1])
        if TRD_start is None:
            TRD_start = date
            continue
        if date < TRD_start:
            TRD_start = date

    # Iterate over rows to process data
    for row in sheet.iter_rows(min_row=2, values_only=True):
        ter = extract_ter(row[2])
        sec = extract_sec(row[2])
        debit = row[3]
        duration = row[11]
        if duration == 0:
            continue
        date_open = None
        if row[9] is not None:
            date_open = extract_date(row[9], row[10])
        if date_open is None:
            continue
        if TRD_end is None:
            TRD_end = date_open + timedelta(hours=duration)
        if date_open + timedelta(hours=duration) > TRD_end:
            TRD_end = date_open + timedelta(hours=duration)
        if sec not in mv2:
            mv2[sec] = {}
        if ter not in mv2[sec]:
            mv2[sec][ter] = {"hours_list": [],
                             "hours_list_summed": [], "total_hours": 0}

        date_diff = calculate_hours_difference(TRD_start, date_open)
        for i in range(duration):
            hour_index = int(date_diff + i)
            if hour_index >= len(mv2[sec][ter]["hours_list"]):
                second_list = [0] * (
                    hour_index - len(mv2[sec][ter]["hours_list"]) + 1
                )
                mv2[sec][ter]["hours_list"].extend(second_list)
                mv2[sec][ter]["hours_list"][hour_index] += debit
            else:
                mv2[sec][ter]["hours_list"][hour_index] += debit

    # Sort the sec and ter keys based on the numeric part
    mv2 = dict(sorted(mv2.items(), key=lambda item: int(
        ''.join(filter(str.isdigit, item[0])))))
    for sec in mv2:
        mv2[sec] = dict(sorted(mv2[sec].items(), key=lambda item: int(
            ''.join(filter(str.isdigit, item[0])))))

    # size of chunks, 15H for night, 9H for day
    CHUNKS_PATTERN = [15, 9]
    for sec in mv2:
        for ter in mv2[sec]:
            summed_list = sum_chunks(
                mv2[sec][ter]["hours_list"], CHUNKS_PATTERN)
            for value in summed_list:
                if value == 0:
                    mv2[sec][ter]["hours_list_summed"].append("")
                else:
                    mv2[sec][ter]["hours_list_summed"].append(
                        round(value, 2))
            mv2[sec][ter]["total_hours"] = sum(
                mv2[sec][ter]["hours_list"]) / 20

    return mv2, TRD_start, TRD_end


def create_mv2(mv2, TRD_start, TRD_end, directory):
    """
    Creates a new Excel file (MV2) from the processed data.

    Args:
        mv2 (dict): The dictionary containing the processed data.
        TRD_start (datetime): The earliest start date.
        directory (str): The directory where the new file will be saved.

    Returns:
        str: The full path of the created Excel file.
    """
    # Calculate the number of days in the TRD period
    TRD_days = (TRD_end - TRD_start).days
    print(f"DEBUG: TRD_days calculated as {TRD_days}")

    # Create a new workbook
    workbook = Workbook()

    # Get the active worksheet
    sheet = workbook.active

    # define grand total row
    grand_total_hours = [0]+[0, 0] * TRD_days

    # Define the style for the border's sides
    thin_black = Side(border_style="thin", color="000000")
    # Define the border using the sides
    border_all = Border(
        left=thin_black, right=thin_black, top=thin_black, bottom=thin_black
    )
    # Define fill
    fill = PatternFill(start_color="DDDDDD",
                       end_color="DDDDDD", fill_type="solid")
    # Define different Alignment objects
    title_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    data_alignment = Alignment(horizontal="center", vertical="center")
    keys_alignment = Alignment(horizontal="left", vertical="center")

    for sec in mv2:
        # Add an empty row between different 'sec' groups
        sheet.append([])
        sheet.append(
            [
                "",
                "Sec",
                "Ter",
                "Duration"
            ]
            + ["N", "J"] * TRD_days
        )
        # formatting header row
        for cell in sheet[sheet.max_row]:
            # Skip the first cell
            if cell.column == 1:
                continue
            cell.font = Font(name="Book Antiqua", size=10, bold=True)
            cell.alignment = title_alignment
            cell.fill = fill
            cell.border = border_all

        sec_total_duration = 0
        for ter in mv2[sec]:
            ter_duration = sum(mv2[sec][ter]["hours_list"]) / 20
            # Update grand total hours
            grand_total_hours[0] += ter_duration
            for i in range(len(mv2[sec][ter]["hours_list_summed"])):
                if isinstance(mv2[sec][ter]["hours_list_summed"][i], (int, float)):
                    grand_total_hours[1 +
                                      i] += mv2[sec][ter]["hours_list_summed"][i]
            sec_total_duration += ter_duration
            sheet.append(
                [""] + [sec, ter, ter_duration] +
                mv2[sec][ter]["hours_list_summed"]
            )
            # formatting data row
            for cell in sheet[sheet.max_row]:
                # Skip the first cell
                if cell.column == 1:
                    continue
                cell.font = Font(name="Book Antiqua", size=10)
                cell.border = border_all
                if cell.column in [2, 3]:  # 'Sec' and 'Ter' columns
                    cell.alignment = keys_alignment
                else:
                    cell.alignment = data_alignment
                if cell.column in [2, 3, 4]:  # 'Sec', 'Ter', and 'Duration' columns
                    cell.font = Font(name="Book Antiqua", size=10, bold=True)

        # Add a total row for the current 'sec'
        total_row = [""] + [sec, "Total", sec_total_duration]
        for i in range(5, 5 + 2 * TRD_days):
            col_sum = sum(
                sheet.cell(row=row_idx, column=i).value or 0
                for row_idx in range(sheet.max_row - len(mv2[sec]) + 1, sheet.max_row + 1)
            )
            total_row.append(round(col_sum, 2) if col_sum != 0 else "")
        sheet.append(total_row)
        # formatting total row
        for cell in sheet[sheet.max_row]:
            # Skip the first cell
            if cell.column == 1:
                continue
            cell.font = Font(name="Book Antiqua", size=10, bold=True)
            cell.alignment = title_alignment
            cell.fill = fill
            cell.border = border_all

    # Add a grand total row at the end
    sheet.append([])
    grand_total_row = ["", "Grand Total", ""] + grand_total_hours
    sheet.append(grand_total_row)
    # formatting grand total row
    for cell in sheet[sheet.max_row]:
        # Skip the first cell
        if cell.column == 1:
            continue
        cell.font = Font(name="Book Antiqua", size=12, bold=True)
        # check if cell value is string
        if isinstance(cell.value, str):
            cell.alignment = keys_alignment
        else:
            cell.alignment = data_alignment

    # add credit note
    sheet.append([])
    sheet.append(
        ["", "Generated by MV2 Creator app - by Anas Asimi - 2025"])
    sheet["B" + str(sheet.max_row)
          ].font = Font(name="Book Antiqua", size=10, italic=True)
    sheet["B" + str(sheet.max_row)].alignment = Alignment(horizontal="left")

    # Save the workbook to a file
    excel_file_name = f"MV2 - {TRD_start.strftime('%Y-%m-%d')}.xlsx"
    excel_full_path = os.path.join(directory, excel_file_name)
    workbook.save(excel_full_path)
    return excel_full_path


# if script is run directly
if __name__ == "__main__":
    # Example usage
    original_path = "mv1.xlsx"
    mv2, TRD_start, TRD_end = xls_to_dict(original_path)
    print(f"TRD_start: {TRD_start}, TRD_end: {TRD_end}")
    print("Data processed successfully.")
    # save mv2 to json for debugging
    import json
    with open("mv2_debug.json", "w") as f:
        json.dump(mv2, f, indent=4)
        print("New MV2 data saved to mv2_debug.json for debugging.")
    directory = os.path.dirname(original_path)
    new_file_path = create_mv2(mv2, TRD_start, TRD_end, directory)
    print(f"New MV2 file created at: {new_file_path}")
